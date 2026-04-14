import tempfile
import unittest
from pathlib import Path

from openpyxl import load_workbook

from upset_model.excel_report import build_prediction_excel_report
from upset_model.modeling import SoftmaxModelArtifact


def make_artifact(labels: list[str]) -> SoftmaxModelArtifact:
    metrics = {
        "accuracy": 0.73,
        "top_10_upset_precision": 0.30,
        "top_20_upset_precision": 0.35,
        "top_20_home_upset_win_precision": 0.25,
        "top_20_away_upset_win_precision": 0.20,
        "top_20_draw_upset_precision": 0.30,
    }
    decision_metrics = {
        "accuracy": 0.70,
        "upset_precision": 0.40,
        "upset_recall": 0.38,
        "draw_precision": 0.31,
        "draw_recall": 0.55,
    }
    return SoftmaxModelArtifact(
        created_at_utc="2026-04-14T00:00:00+00:00",
        labels=labels,
        feature_names=["feature_a"],
        feature_means=[0.0],
        feature_stds=[1.0],
        weights=[[0.0, 0.0] for _ in labels],
        class_weights={label: 1.0 for label in labels},
        validation_season="2526",
        train_size=100,
        validation_size=50,
        metrics=metrics,
        decision_threshold=0.62,
        decision_metrics=decision_metrics,
    )


class ExcelReportTests(unittest.TestCase):
    def test_build_prediction_excel_report_creates_expected_sheets(self) -> None:
        run_summary = {
            "run_id": "20260414T000000Z",
            "start_date": "2026-04-18",
            "end_date": "2026-04-20",
            "competition_filter_mode": "all",
            "requested_competitions": [],
            "selected_competitions": ["E0", "I1"],
            "scheduled_match_count": 10,
            "scored_row_count": 10,
            "failure_count": 0,
            "prediction_csv_path": "/tmp/predictions.csv",
            "combined_ranking_csv_path": "/tmp/combined.csv",
            "draw_ranking_csv_path": "/tmp/draw.csv",
            "betting_ranking_csv_path": "/tmp/betting.csv",
        }
        combined_rows = [
            {
                "match_date": "2026-04-18",
                "kickoff_time": "15:00",
                "competition_code": "E0",
                "competition_name": "Premier League",
                "home_team": "A",
                "away_team": "B",
                "combined_candidate_label": "draw_upset",
                "combined_candidate_probability": "0.56",
                "secondary_candidate_label": "home_upset_win",
                "secondary_candidate_probability": "0.31",
                "home_upset_probability": "0.31",
                "draw_upset_probability": "0.56",
                "away_upset_probability": "0.22",
                "upset_score": "0.53",
                "predicted_label": "non_upset",
                "explanation": "x",
            }
        ]
        recommendation_rows = [
            {
                "比赛时间": "2026-04-18 15:00",
                "对阵": "A vs B",
                "等级": "强",
                "建议方向": "冷平",
                "原因": "冷平概率 0.56，达到门槛 0.48",
            }
        ]
        all_rows = [
            {
                **combined_rows[0],
                "candidate_label": "home_upset_win",
                "candidate_probability": "0.31",
                "non_upset_probability": "0.47",
                "actual_label": "unknown",
                "bet_direction": "冷平",
                "bet_confidence": "强",
                "bet_recommendation": "建议投注",
                "direction_probability": "0.56",
                "direction_gap": "0.25",
                "bet_reason": "冷平概率 0.56，达到门槛 0.48",
            }
        ]
        with tempfile.TemporaryDirectory() as tmp_dir:
            output_path = Path(tmp_dir) / "report.xlsx"
            build_prediction_excel_report(
                run_summary=run_summary,
                combined_rows=combined_rows,
                recommendation_rows=recommendation_rows,
                draw_rows=combined_rows,
                all_rows=all_rows,
                main_artifact=make_artifact(["home_upset_win", "away_upset_win", "non_upset"]),
                draw_artifact=make_artifact(["draw_upset", "non_draw_upset"]),
                output_path=output_path,
            )
            workbook = load_workbook(output_path)

        self.assertEqual(
            workbook.sheetnames,
            ["summary", "betting_recommendations", "combined_rankings", "draw_rankings", "all_predictions"],
        )
        self.assertEqual(workbook["summary"]["A1"].value, "Titan007 Prediction Summary")
        self.assertEqual(workbook["betting_recommendations"]["C2"].value, "强")
        self.assertEqual(workbook["betting_recommendations"]["D2"].value, "冷平")
        self.assertEqual(workbook["combined_rankings"]["G2"].value, "draw_upset")
        self.assertAlmostEqual(workbook["combined_rankings"]["H2"].value, 0.56)
        summary_values = [cell for row in workbook["summary"].iter_rows(values_only=True) for cell in row if cell is not None]
        self.assertIn("比赛时间", summary_values)
        self.assertIn("等级", summary_values)
        self.assertIn("建议方向", summary_values)

    def test_build_prediction_excel_report_keeps_draw_sheet_when_model_missing(self) -> None:
        with tempfile.TemporaryDirectory() as tmp_dir:
            output_path = Path(tmp_dir) / "report.xlsx"
            build_prediction_excel_report(
                run_summary={"run_id": "x"},
                combined_rows=[],
                recommendation_rows=[],
                draw_rows=[],
                all_rows=[],
                main_artifact=make_artifact(["home_upset_win", "away_upset_win", "non_upset"]),
                draw_artifact=None,
                output_path=output_path,
            )
            workbook = load_workbook(output_path)

        self.assertIn("draw_rankings", workbook.sheetnames)
        self.assertIn("Draw model unavailable", str(workbook["draw_rankings"]["A1"].value))


if __name__ == "__main__":
    unittest.main()

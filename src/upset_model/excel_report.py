from __future__ import annotations

import csv
import json
from pathlib import Path
from typing import Iterable

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from upset_model.modeling import SoftmaxModelArtifact

LABEL_FILLS = {
    "home_upset_win": PatternFill(fill_type="solid", fgColor="DDEBF7"),
    "draw_upset": PatternFill(fill_type="solid", fgColor="FFF2CC"),
    "away_upset_win": PatternFill(fill_type="solid", fgColor="FCE4D6"),
}
BET_DIRECTION_FILLS = {
    "主冷": PatternFill(fill_type="solid", fgColor="DDEBF7"),
    "冷平": PatternFill(fill_type="solid", fgColor="FFF2CC"),
    "客冷": PatternFill(fill_type="solid", fgColor="FCE4D6"),
    "不投注": PatternFill(fill_type="solid", fgColor="E7E6E6"),
}
BET_CONFIDENCE_FILLS = {
    "强": PatternFill(fill_type="solid", fgColor="C6E0B4"),
    "中": PatternFill(fill_type="solid", fgColor="FFE699"),
    "弱": PatternFill(fill_type="solid", fgColor="F4B183"),
    "不投注": PatternFill(fill_type="solid", fgColor="E7E6E6"),
}
HEADER_FILL = PatternFill(fill_type="solid", fgColor="1F4E78")
HEADER_FONT = Font(color="FFFFFF", bold=True)
SECTION_FONT = Font(bold=True)

PERCENT_COLUMNS = {
    "combined_candidate_probability",
    "secondary_candidate_probability",
    "home_upset_probability",
    "draw_upset_probability",
    "away_upset_probability",
    "upset_score",
    "candidate_probability",
    "direction_probability",
    "direction_gap",
}


def _load_csv_rows(path: Path) -> list[dict[str, str]]:
    if not path.exists():
        return []
    with path.open(encoding="utf-8", newline="") as handle:
        return list(csv.DictReader(handle))


def _load_json(path: Path) -> dict[str, object]:
    return json.loads(path.read_text(encoding="utf-8"))


def _coerce_cell_value(column_name: str, raw_value: str | None) -> object:
    if raw_value is None or raw_value == "":
        return ""
    if column_name in PERCENT_COLUMNS:
        try:
            return float(raw_value)
        except ValueError:
            return raw_value
    return raw_value


def _style_header_row(worksheet, header_row: int = 1) -> None:
    for cell in worksheet[header_row]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")


def _autosize_columns(worksheet) -> None:
    for column_cells in worksheet.columns:
        max_length = 0
        column_letter = get_column_letter(column_cells[0].column)
        header_value = "" if column_cells[0].value is None else str(column_cells[0].value)
        for cell in column_cells:
            value = "" if cell.value is None else str(cell.value)
            max_length = max(max_length, len(value))
        max_width = 100 if header_value == "原因" else 40
        worksheet.column_dimensions[column_letter].width = min(max(max_length + 2, 10), max_width)


def _apply_percent_format(worksheet, header_names: list[str], start_row: int, end_row: int) -> None:
    for column_index, header_name in enumerate(header_names, start=1):
        if header_name not in PERCENT_COLUMNS:
            continue
        for row_index in range(start_row, end_row + 1):
            worksheet.cell(row=row_index, column=column_index).number_format = "0.00%"


def _apply_label_fill(worksheet, header_names: list[str], start_row: int, end_row: int) -> None:
    fill_mappings = {
        "combined_candidate_label": LABEL_FILLS,
        "secondary_candidate_label": LABEL_FILLS,
        "bet_direction": BET_DIRECTION_FILLS,
        "bet_confidence": BET_CONFIDENCE_FILLS,
    }
    for target_column, fill_mapping in fill_mappings.items():
        if target_column not in header_names:
            continue
        column_index = header_names.index(target_column) + 1
        for row_index in range(start_row, end_row + 1):
            cell = worksheet.cell(row=row_index, column=column_index)
            fill = fill_mapping.get(str(cell.value))
            if fill is not None:
                cell.fill = fill


def _write_table_sheet(workbook: Workbook, title: str, rows: list[dict[str, str]], empty_message: str) -> None:
    worksheet = workbook.create_sheet(title=title)
    if not rows:
        worksheet["A1"] = empty_message
        worksheet["A1"].font = SECTION_FONT
        _autosize_columns(worksheet)
        return

    header_names = list(rows[0].keys())
    worksheet.append(header_names)
    for row in rows:
        worksheet.append([_coerce_cell_value(column_name, row.get(column_name)) for column_name in header_names])

    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = worksheet.dimensions
    _style_header_row(worksheet)
    for row_index in range(2, worksheet.max_row + 1):
        for column_index in range(1, len(header_names) + 1):
            worksheet.cell(row=row_index, column=column_index).alignment = Alignment(vertical="top", wrap_text=True)
    _apply_percent_format(worksheet, header_names, start_row=2, end_row=worksheet.max_row)
    _apply_label_fill(worksheet, header_names, start_row=2, end_row=worksheet.max_row)
    _autosize_columns(worksheet)


def _write_summary_sheet(
    workbook: Workbook,
    *,
    run_summary: dict[str, object],
    combined_rows: list[dict[str, str]],
    recommendation_rows: list[dict[str, str]],
    main_artifact: SoftmaxModelArtifact,
    draw_artifact: SoftmaxModelArtifact | None,
) -> None:
    worksheet = workbook.active
    worksheet.title = "summary"
    worksheet["A1"] = "Titan007 Prediction Summary"
    worksheet["A1"].font = Font(bold=True, size=14)

    summary_pairs = [
        ("report_generated_utc", run_summary.get("run_id", "")),
        ("start_date", run_summary.get("start_date", "")),
        ("end_date", run_summary.get("end_date", "")),
        ("start_datetime", run_summary.get("start_datetime", "")),
        ("end_datetime", run_summary.get("end_datetime", "")),
        ("competition_filter_mode", run_summary.get("competition_filter_mode", "")),
        ("requested_competitions", ", ".join(run_summary.get("requested_competitions", []))),
        ("selected_competitions", ", ".join(run_summary.get("selected_competitions", []))),
        ("scheduled_match_count", run_summary.get("scheduled_match_count", "")),
        ("scored_row_count", run_summary.get("scored_row_count", "")),
        ("failure_count", run_summary.get("failure_count", "")),
        ("prediction_csv_path", run_summary.get("prediction_csv_path", "")),
        ("combined_ranking_csv_path", run_summary.get("combined_ranking_csv_path", "")),
        ("draw_ranking_csv_path", run_summary.get("draw_ranking_csv_path", "")),
        ("betting_ranking_csv_path", run_summary.get("betting_ranking_csv_path", "")),
        ("betting_ranking_json_path", run_summary.get("betting_ranking_json_path", "")),
    ]

    row_index = 3
    worksheet.cell(row=row_index, column=1, value="run_summary").font = SECTION_FONT
    row_index += 1
    for key, value in summary_pairs:
        worksheet.cell(row=row_index, column=1, value=key)
        worksheet.cell(row=row_index, column=2, value=value)
        row_index += 1

    row_index += 1
    worksheet.cell(row=row_index, column=1, value="main_model_metrics").font = SECTION_FONT
    row_index += 1
    main_metrics = [
        ("validation_season", main_artifact.validation_season),
        ("accuracy", main_artifact.metrics.get("accuracy")),
        ("top_10_upset_precision", main_artifact.metrics.get("top_10_upset_precision")),
        ("top_20_upset_precision", main_artifact.metrics.get("top_20_upset_precision")),
        ("top_20_home_upset_win_precision", main_artifact.metrics.get("top_20_home_upset_win_precision")),
        ("top_20_away_upset_win_precision", main_artifact.metrics.get("top_20_away_upset_win_precision")),
        ("decision_threshold", main_artifact.decision_threshold),
        ("decision_accuracy", (main_artifact.decision_metrics or {}).get("accuracy")),
        ("decision_upset_precision", (main_artifact.decision_metrics or {}).get("upset_precision")),
        ("decision_upset_recall", (main_artifact.decision_metrics or {}).get("upset_recall")),
    ]
    for key, value in main_metrics:
        worksheet.cell(row=row_index, column=1, value=key)
        worksheet.cell(row=row_index, column=2, value=value)
        row_index += 1

    row_index += 1
    worksheet.cell(row=row_index, column=1, value="draw_model_metrics").font = SECTION_FONT
    row_index += 1
    if draw_artifact is None:
        worksheet.cell(row=row_index, column=1, value="draw_model_status")
        worksheet.cell(row=row_index, column=2, value="not_loaded")
        row_index += 1
    else:
        draw_metrics = [
            ("validation_season", draw_artifact.validation_season),
            ("accuracy", draw_artifact.metrics.get("accuracy")),
            ("top_20_draw_upset_precision", draw_artifact.metrics.get("top_20_draw_upset_precision")),
            ("decision_threshold", draw_artifact.decision_threshold),
            ("decision_draw_precision", (draw_artifact.decision_metrics or {}).get("draw_precision")),
            ("decision_draw_recall", (draw_artifact.decision_metrics or {}).get("draw_recall")),
        ]
        for key, value in draw_metrics:
            worksheet.cell(row=row_index, column=1, value=key)
            worksheet.cell(row=row_index, column=2, value=value)
            row_index += 1

    row_index += 1
    worksheet.cell(row=row_index, column=1, value="combined_top_preview").font = SECTION_FONT
    row_index += 1
    preview_headers = [
        "match_date",
        "competition_code",
        "home_team",
        "away_team",
        "combined_candidate_label",
        "combined_candidate_probability",
        "secondary_candidate_label",
        "secondary_candidate_probability",
    ]
    for column_index, header_name in enumerate(preview_headers, start=1):
        worksheet.cell(row=row_index, column=column_index, value=header_name)
    _style_header_row(worksheet, header_row=row_index)
    preview_header_row = row_index
    row_index += 1
    for row in combined_rows[:10]:
        for column_index, header_name in enumerate(preview_headers, start=1):
            worksheet.cell(
                row=row_index,
                column=column_index,
                value=_coerce_cell_value(header_name, row.get(header_name)),
            )
        row_index += 1

    _apply_percent_format(worksheet, preview_headers, start_row=preview_header_row + 1, end_row=max(preview_header_row + 1, row_index - 1))
    _apply_label_fill(worksheet, preview_headers, start_row=preview_header_row + 1, end_row=max(preview_header_row + 1, row_index - 1))

    row_index += 1
    worksheet.cell(row=row_index, column=1, value="betting_top_preview").font = SECTION_FONT
    row_index += 1
    betting_headers = [
        "比赛时间",
        "对阵",
        "等级",
        "建议方向",
        "原因",
    ]
    for column_index, header_name in enumerate(betting_headers, start=1):
        worksheet.cell(row=row_index, column=column_index, value=header_name)
    _style_header_row(worksheet, header_row=row_index)
    betting_header_row = row_index
    row_index += 1
    for row in recommendation_rows[:10]:
        for column_index, header_name in enumerate(betting_headers, start=1):
            worksheet.cell(
                row=row_index,
                column=column_index,
                value=_coerce_cell_value(header_name, row.get(header_name)),
            )
        row_index += 1

    _apply_percent_format(worksheet, betting_headers, start_row=betting_header_row + 1, end_row=max(betting_header_row + 1, row_index - 1))
    _apply_label_fill(worksheet, betting_headers, start_row=betting_header_row + 1, end_row=max(betting_header_row + 1, row_index - 1))
    _autosize_columns(worksheet)


def build_prediction_excel_report(
    *,
    run_summary: dict[str, object],
    combined_rows: list[dict[str, str]],
    recommendation_rows: list[dict[str, str]],
    draw_rows: list[dict[str, str]],
    all_rows: list[dict[str, str]],
    main_artifact: SoftmaxModelArtifact,
    draw_artifact: SoftmaxModelArtifact | None,
    output_path: Path,
) -> Path:
    workbook = Workbook()
    _write_summary_sheet(
        workbook,
        run_summary=run_summary,
        combined_rows=combined_rows,
        recommendation_rows=recommendation_rows,
        main_artifact=main_artifact,
        draw_artifact=draw_artifact,
    )
    _write_table_sheet(
        workbook,
        "betting_recommendations",
        recommendation_rows,
        empty_message="No betting recommendation rows available.",
    )
    _write_table_sheet(workbook, "combined_rankings", combined_rows, empty_message="No combined ranking rows available.")
    _write_table_sheet(
        workbook,
        "draw_rankings",
        draw_rows,
        empty_message="Draw model unavailable or no draw ranking rows available.",
    )
    _write_table_sheet(workbook, "all_predictions", all_rows, empty_message="No prediction rows available.")
    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)
    return output_path


def build_prediction_excel_report_from_run_dir(
    *,
    run_dir: Path,
    main_artifact: SoftmaxModelArtifact,
    draw_artifact: SoftmaxModelArtifact | None,
    output_path: Path | None = None,
) -> Path:
    run_summary = _load_json(run_dir / "run_summary.json")
    combined_rows = _load_csv_rows(run_dir / "combined_rankings.csv")
    recommendation_rows = _load_csv_rows(run_dir / "betting_recommendations.csv")
    draw_rows = _load_csv_rows(run_dir / "draw_rankings.csv")
    all_rows = _load_csv_rows(run_dir / "predictions.csv")
    target = output_path or (run_dir / "prediction_report.xlsx")
    return build_prediction_excel_report(
        run_summary=run_summary,
        combined_rows=combined_rows,
        recommendation_rows=recommendation_rows,
        draw_rows=draw_rows,
        all_rows=all_rows,
        main_artifact=main_artifact,
        draw_artifact=draw_artifact,
        output_path=target,
    )
